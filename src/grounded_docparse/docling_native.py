from __future__ import annotations

import csv
import hashlib
import posixpath
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from typing import Any
from zipfile import ZipFile

from defusedxml import ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .native import (
    CellSourceAnchor,
    CsvSourceAnchor,
    NativeAsset,
    PageRoute,
    SourceAnchor,
    SourceFormat,
    SourceUnit,
    StructuralSourceAnchor,
    TextSourceAnchor,
)


@dataclass(slots=True)
class SourceRecord:
    text: str
    type: str
    anchor: SourceAnchor
    cells: list[tuple[int, int, str]] | None = None
    claimed: bool = False


@dataclass(frozen=True, slots=True)
class SourceManifest:
    units: list[SourceUnit]
    records: list[SourceRecord]
    assets: list[NativeAsset]


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _text(node: Any) -> str:
    return "".join(value or "" for value in node.itertext()).strip()


def _norm(value: str) -> str:
    return " ".join(value.split()).casefold()


def claim_record(
    records: list[SourceRecord],
    text: str,
    *,
    unit_id: str | None = None,
    record_type: str | None = None,
) -> SourceRecord:
    normalized = _norm(text)
    matches = [
        record
        for record in records
        if not record.claimed
        and _norm(record.text) == normalized
        and (unit_id is None or record.anchor.unit_id == unit_id)
        and (record_type is None or record.type == record_type)
    ]
    if not matches and record_type == "table":
        matches = [
            record
            for record in records
            if not record.claimed
            and record.type == "table"
            and (unit_id is None or record.anchor.unit_id == unit_id)
        ]
        if len(matches) != 1:
            matches = []
    if not matches:
        raise ValueError(f"Docling block cannot be mapped to its source: {text[:80]!r}")
    record = matches[0]
    record.claimed = True
    return record


def _unit(unit_id: str, kind: str, index: int, label: str | None = None) -> SourceUnit:
    return SourceUnit(
        id=unit_id,
        kind=kind,
        index=index,
        label=label,
        requested_route=PageRoute.NATIVE,
        effective_route=PageRoute.NATIVE,
        parser="docling",
    )


def _media_type(name: str) -> str | None:
    suffix = PurePosixPath(name).suffix.casefold()
    return {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
        ".tif": "image/tiff",
        ".tiff": "image/tiff",
        ".svg": "image/svg+xml",
        ".webp": "image/webp",
    }.get(suffix)


def _package_assets(
    archive: ZipFile, prefix: str, unit_id: str
) -> list[NativeAsset]:
    assets = []
    for index, name in enumerate(
        sorted(item for item in archive.namelist() if item.startswith(prefix)), start=1
    ):
        content = archive.read(name)
        assets.append(
            NativeAsset(
                id=f"asset-{index}",
                anchor=StructuralSourceAnchor(unit_id=unit_id, path=f"/{name}"),
                media_type=_media_type(name),
                filename=PurePosixPath(name).name,
                reference=name,
                sha256=hashlib.sha256(content).hexdigest(),
            )
        )
    return assets


def _docx(data: bytes) -> SourceManifest:
    unit = _unit("document-1", "document", 1)
    records: list[SourceRecord] = []
    with ZipFile(BytesIO(data)) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
        body = next(node for node in root.iter() if _local(node.tag) == "body")
        paragraph = table = 0
        for child in body:
            if _local(child.tag) == "p":
                paragraph += 1
                value = _text(child)
                if value:
                    records.append(
                        SourceRecord(
                            value,
                            "paragraph",
                            StructuralSourceAnchor(
                                unit_id=unit.id,
                                path=f"/word/document/body/p[{paragraph}]",
                            ),
                        )
                    )
            elif _local(child.tag) == "tbl":
                table += 1
                cells = [
                    _text(node)
                    for node in child.iter()
                    if _local(node.tag) == "tc"
                ]
                records.append(
                    SourceRecord(
                        " ".join(cells),
                        "table",
                        StructuralSourceAnchor(
                            unit_id=unit.id,
                            path=f"/word/document/body/tbl[{table}]",
                        ),
                    )
                )
        assets = _package_assets(archive, "word/media/", unit.id)
    return SourceManifest([unit], records, assets)


def _pptx(data: bytes) -> SourceManifest:
    records: list[SourceRecord] = []
    units: list[SourceUnit] = []
    assets: list[NativeAsset] = []
    with ZipFile(BytesIO(data)) as archive:
        slide_names = sorted(
            (
                name
                for name in archive.namelist()
                if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
            ),
            key=lambda name: int(re.search(r"\d+", PurePosixPath(name).stem).group()),
        )
        for slide_index, name in enumerate(slide_names, start=1):
            unit = _unit(f"slide-{slide_index}", "slide", slide_index, f"Slide {slide_index}")
            units.append(unit)
            root = ET.fromstring(archive.read(name))
            shape_order = 0
            for node in root.iter():
                if _local(node.tag) not in {"sp", "graphicFrame"}:
                    continue
                shape_order += 1
                id_node = next(
                    (part for part in node.iter() if _local(part.tag) == "cNvPr"), None
                )
                shape_id = (
                    id_node.attrib.get("id", str(shape_order))
                    if id_node is not None
                    else str(shape_order)
                )
                table_node = next(
                    (part for part in node.iter() if _local(part.tag) == "tbl"), None
                )
                if table_node is not None:
                    value = _text(table_node)
                    if not value:
                        continue
                    records.append(
                        SourceRecord(
                            value,
                            "table",
                            StructuralSourceAnchor(
                                unit_id=unit.id,
                                path=f"/ppt/slides/slide[{slide_index}]/shape[id={shape_id}]",
                            ),
                        )
                    )
                    continue
                paragraph_index = 0
                for paragraph_node in (
                    part for part in node.iter() if _local(part.tag) == "p"
                ):
                    paragraph_index += 1
                    value = _text(paragraph_node)
                    if not value:
                        continue
                    records.append(
                        SourceRecord(
                            value,
                            "text",
                            StructuralSourceAnchor(
                                unit_id=unit.id,
                                path=(
                                    f"/ppt/slides/slide[{slide_index}]"
                                    f"/shape[id={shape_id}]/paragraph[{paragraph_index}]"
                                ),
                            ),
                        )
                    )
        for asset in _package_assets(archive, "ppt/media/", units[0].id):
            assets.append(asset.model_copy(update={"id": f"asset-{len(assets) + 1}"}))
    return SourceManifest(units, records, assets)


def _xlsx(data: bytes) -> SourceManifest:
    workbook = load_workbook(BytesIO(data), read_only=False, data_only=False)
    records: list[SourceRecord] = []
    units: list[SourceUnit] = []
    assets: list[NativeAsset] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            unit = _unit(f"sheet-{sheet_index}", "sheet", sheet_index, sheet.title)
            units.append(unit)
            populated = [cell for row in sheet.iter_rows() for cell in row if cell.value is not None]
            if populated:
                min_row = min(cell.row for cell in populated)
                max_row = max(cell.row for cell in populated)
                min_col = min(cell.column for cell in populated)
                max_col = max(cell.column for cell in populated)
                values = [str(cell.value) for cell in populated]
                records.append(
                    SourceRecord(
                        " ".join(values),
                        "table",
                        CellSourceAnchor(
                            unit_id=unit.id,
                            sheet=sheet.title,
                            cell_range=(
                                f"{get_column_letter(min_col)}{min_row}:"
                                f"{get_column_letter(max_col)}{max_row}"
                            ),
                        ),
                        cells=[
                            (cell.row - min_row, cell.column - min_col, str(cell.value))
                            for cell in populated
                        ],
                    )
                )
            for image_index, image in enumerate(getattr(sheet, "_images", []), start=1):
                anchor = getattr(image, "anchor", None)
                marker = getattr(anchor, "_from", None)
                cell = (
                    f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
                    if marker is not None
                    else "A1"
                )
                raw = image._data()
                assets.append(
                    NativeAsset(
                        id=f"asset-{len(assets) + 1}",
                        anchor=CellSourceAnchor(
                            unit_id=unit.id, sheet=sheet.title, cell_range=cell
                        ),
                        media_type=f"image/{image.format}" if image.format else None,
                        filename=f"{sheet.title}-image-{image_index}.{image.format or 'bin'}",
                        sha256=hashlib.sha256(raw).hexdigest(),
                        width=int(image.width) if image.width else None,
                        height=int(image.height) if image.height else None,
                    )
                )
    finally:
        workbook.close()
    return SourceManifest(units, records, assets)


def _csv_manifest(data: bytes) -> SourceManifest:
    rows = list(csv.reader(StringIO(data.decode("utf-8-sig"))))
    unit = _unit("document-1", "document", 1)
    width = max((len(row) for row in rows), default=1)
    records = []
    if rows:
        records.append(
            SourceRecord(
                " ".join(cell for row in rows for cell in row),
                "table",
                CsvSourceAnchor(
                    unit_id=unit.id,
                    row_start=1,
                    row_end=len(rows),
                    column_start=1,
                    column_end=width,
                ),
            )
        )
    return SourceManifest([unit], records, [])


def _html_records(
    content: str, *, unit: SourceUnit, prefix: str = ""
) -> tuple[list[SourceRecord], list[NativeAsset]]:
    try:
        from bs4 import BeautifulSoup, Tag
    except ImportError as exc:
        raise RuntimeError("native HTML parsing requires grounded-docparse[native]") from exc
    soup = BeautifulSoup(content, "html.parser")
    block_names = {"p", "h1", "h2", "h3", "h4", "h5", "h6", "li", "pre", "table"}

    def path(node: Tag) -> str:
        parts = []
        current: Tag | None = node
        while current is not None and current.name != "[document]":
            siblings = [item for item in current.parent.children if isinstance(item, Tag) and item.name == current.name]
            parts.append(f"{current.name}[{siblings.index(current) + 1}]")
            current = current.parent if isinstance(current.parent, Tag) else None
        return "/" + "/".join(reversed(parts))

    records = []
    for node in soup.find_all(block_names):
        if node.name != "table" and node.find(block_names):
            continue
        value = node.get_text(" ", strip=True)
        if value:
            records.append(
                SourceRecord(
                    value,
                    "table" if node.name == "table" else "text",
                    StructuralSourceAnchor(
                        unit_id=unit.id, path=f"{prefix}{path(node)}"
                    ),
                )
            )
    assets = []
    for index, image in enumerate(soup.find_all("img"), start=1):
        reference = image.get("src")
        assets.append(
            NativeAsset(
                id=f"asset-{index}",
                anchor=StructuralSourceAnchor(
                    unit_id=unit.id, path=f"{prefix}{path(image)}"
                ),
                reference=reference,
                filename=PurePosixPath(reference).name if reference else None,
                media_type=_media_type(reference or ""),
                alt_text=image.get("alt"),
            )
        )
    return records, assets


def _html(data: bytes) -> SourceManifest:
    unit = _unit("document-1", "document", 1)
    records, assets = _html_records(data.decode("utf-8-sig"), unit=unit)
    return SourceManifest([unit], records, assets)


def _epub(data: bytes) -> SourceManifest:
    units: list[SourceUnit] = []
    records: list[SourceRecord] = []
    assets: list[NativeAsset] = []
    with ZipFile(BytesIO(data)) as archive:
        container = ET.fromstring(archive.read("META-INF/container.xml"))
        rootfile = next(node for node in container.iter() if _local(node.tag) == "rootfile")
        package_name = rootfile.attrib["full-path"]
        package = ET.fromstring(archive.read(package_name))
        manifest = {
            node.attrib["id"]: node.attrib["href"]
            for node in package.iter()
            if _local(node.tag) == "item" and "id" in node.attrib
        }
        spine = [
            node.attrib["idref"]
            for node in package.iter()
            if _local(node.tag) == "itemref"
        ]
        base = posixpath.dirname(package_name)
        for index, item_id in enumerate(spine, start=1):
            href = manifest[item_id]
            name = posixpath.normpath(posixpath.join(base, href))
            unit = _unit(f"section-{index}", "section", index, href)
            units.append(unit)
            item_records, item_assets = _html_records(
                archive.read(name).decode("utf-8"), unit=unit, prefix=f"epub:{href}:"
            )
            records.extend(item_records)
            for asset in item_assets:
                reference = asset.reference
                if reference and not re.match(
                    r"^[a-z][a-z0-9+.-]*:", reference, re.IGNORECASE
                ):
                    target = posixpath.normpath(posixpath.join(posixpath.dirname(name), reference))
                    if target in archive.namelist():
                        content = archive.read(target)
                        asset = asset.model_copy(
                            update={
                                "reference": target,
                                "sha256": hashlib.sha256(content).hexdigest(),
                                "media_type": _media_type(target),
                            }
                        )
                assets.append(asset.model_copy(update={"id": f"asset-{len(assets) + 1}"}))
    return SourceManifest(units, records, assets)


def _markdown(data: bytes) -> SourceManifest:
    value = data.decode("utf-8-sig")
    lines = value.splitlines()
    unit = _unit("document-1", "document", 1)
    records = []
    start = 0
    for index in range(len(lines) + 1):
        if index < len(lines) and lines[index].strip():
            continue
        if start < index:
            block = "\n".join(lines[start:index])
            is_table = all(line.lstrip().startswith("|") for line in lines[start:index])
            if is_table:
                table_rows = [
                    [cell.strip() for cell in line.strip().strip("|").split("|")]
                    for line in lines[start:index]
                    if not re.fullmatch(r"\s*\|?(?:\s*:?-+:?\s*\|)+\s*", line)
                ]
                plain = " ".join(cell for row in table_rows for cell in row)
                anchor: SourceAnchor = StructuralSourceAnchor(
                    unit_id=unit.id, path=f"/markdown/lines[{start + 1}:{index}]"
                )
                record_type = "table"
            else:
                plain = re.sub(r"^(?:#{1,6}|[-*+] |\d+[.)] )\s*", "", block).strip()
                anchor = TextSourceAnchor(
                    unit_id=unit.id,
                    start_line=start + 1,
                    end_line=index,
                    start_column=1,
                    end_column=len(lines[index - 1]) + 1,
                )
                record_type = "text"
            records.append(
                SourceRecord(
                    plain,
                    record_type,
                    anchor,
                )
            )
        start = index + 1
    return SourceManifest([unit], records, [])


def _odf(data: bytes, source_format: SourceFormat) -> SourceManifest:
    with ZipFile(BytesIO(data)) as archive:
        root = ET.fromstring(archive.read("content.xml"))
        records: list[SourceRecord] = []
        units: list[SourceUnit] = []
        if source_format is SourceFormat.ODP:
            pages = [node for node in root.iter() if _local(node.tag) == "page"]
            for page_index, page in enumerate(pages, start=1):
                page_name = next(
                    (
                        value
                        for key, value in page.attrib.items()
                        if _local(key) == "name"
                    ),
                    f"Slide {page_index}",
                )
                unit = _unit(f"slide-{page_index}", "slide", page_index, page_name)
                units.append(unit)
                for shape_index, node in enumerate(
                    [part for part in page.iter() if _local(part.tag) in {"frame", "custom-shape"}],
                    start=1,
                ):
                    value = _text(node)
                    if value:
                        records.append(
                            SourceRecord(
                                value,
                                "text",
                                StructuralSourceAnchor(
                                    unit_id=unit.id,
                                    path=f"/office:presentation/draw:page[{page_index}]/shape[{shape_index}]",
                                ),
                            )
                        )
        elif source_format is SourceFormat.ODS:
            sheets = [node for node in root.iter() if _local(node.tag) == "table"]
            for sheet_index, sheet in enumerate(sheets, start=1):
                name = next((value for key, value in sheet.attrib.items() if _local(key) == "name"), f"Sheet {sheet_index}")
                unit = _unit(f"sheet-{sheet_index}", "sheet", sheet_index, name)
                units.append(unit)
                rows: list[list[str]] = []
                for row in (node for node in sheet.iter() if _local(node.tag) == "table-row"):
                    values: list[str] = []
                    for cell in (node for node in row if _local(node.tag) == "table-cell"):
                        repeat = int(next((value for key, value in cell.attrib.items() if _local(key) == "number-columns-repeated"), "1"))
                        values.extend([_text(cell)] * repeat)
                    rows.append(values)
                populated = [
                    (row_index, column_index, value)
                    for row_index, row in enumerate(rows, start=1)
                    for column_index, value in enumerate(row, start=1)
                    if value
                ]
                if populated:
                    min_row = min(row for row, _column, _value in populated)
                    max_row = max(row for row, _column, _value in populated)
                    min_column = min(column for _row, column, _value in populated)
                    max_column = max(column for _row, column, _value in populated)
                    records.append(
                        SourceRecord(
                            " ".join(value for _row, _column, value in populated),
                            "table",
                            CellSourceAnchor(
                                unit_id=unit.id,
                                sheet=name,
                                cell_range=(
                                    f"{get_column_letter(min_column)}{min_row}:"
                                    f"{get_column_letter(max_column)}{max_row}"
                                ),
                            ),
                        )
                    )
        else:
            unit = _unit("document-1", "document", 1)
            units.append(unit)
            tables = [node for node in root.iter() if _local(node.tag) == "table"]
            table_descendants = {id(node) for table in tables for node in table.iter()}
            for table_index, table in enumerate(tables, start=1):
                value = _text(table)
                if value:
                    records.append(
                        SourceRecord(
                            value,
                            "table",
                            StructuralSourceAnchor(
                                unit_id=unit.id,
                                path=f"/office:text/table[{table_index}]",
                            ),
                        )
                    )
            paragraph = 0
            for node in root.iter():
                if _local(node.tag) not in {"p", "h"} or id(node) in table_descendants:
                    continue
                paragraph += 1
                value = _text(node)
                if value:
                    records.append(
                        SourceRecord(
                            value,
                            "paragraph",
                            StructuralSourceAnchor(
                                unit_id=unit.id,
                                path=f"/office:text/paragraph[{paragraph}]",
                            ),
                        )
                    )
        assets = _package_assets(archive, "Pictures/", units[0].id)
    return SourceManifest(units, records, assets)


def build_source_manifest(data: bytes, source_format: SourceFormat) -> SourceManifest:
    builders = {
        SourceFormat.DOCX: _docx,
        SourceFormat.PPTX: _pptx,
        SourceFormat.XLSX: _xlsx,
        SourceFormat.CSV: _csv_manifest,
        SourceFormat.HTML: _html,
        SourceFormat.EPUB: _epub,
        SourceFormat.MARKDOWN: _markdown,
    }
    if source_format in {SourceFormat.ODT, SourceFormat.ODP, SourceFormat.ODS}:
        return _odf(data, source_format)
    return builders[source_format](data)


DOCLING_FORMAT_NAMES = {
    SourceFormat.DOCX: "DOCX",
    SourceFormat.PPTX: "PPTX",
    SourceFormat.XLSX: "XLSX",
    SourceFormat.CSV: "CSV",
    SourceFormat.HTML: "HTML",
    SourceFormat.EPUB: "EPUB",
    SourceFormat.ODT: "ODT",
    SourceFormat.ODP: "ODP",
    SourceFormat.ODS: "ODS",
    SourceFormat.MARKDOWN: "MD",
}


def make_docling_converter():
    try:
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import ConvertPipelineOptions
        from docling.document_converter import DocumentConverter
        from docling.pipeline.simple_pipeline import SimplePipeline
    except ImportError as exc:
        raise RuntimeError("native document parsing requires grounded-docparse[native]") from exc

    allowed = [getattr(InputFormat, name) for name in DOCLING_FORMAT_NAMES.values()]
    seed = DocumentConverter(allowed_formats=allowed)
    pipeline_options = ConvertPipelineOptions(
        enable_remote_services=False,
        allow_external_plugins=False,
        do_picture_classification=False,
        do_picture_description=False,
        do_chart_extraction=False,
    )
    format_options = {
        source_format: seed.format_to_options[source_format].model_copy(
            update={
                "pipeline_cls": SimplePipeline,
                "pipeline_options": pipeline_options.model_copy(deep=True),
            }
        )
        for source_format in allowed
    }
    return DocumentConverter(allowed_formats=allowed, format_options=format_options)
